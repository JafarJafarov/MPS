from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from django.shortcuts import render
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .forms import (
    SAPImportForm,
    WeeklyPlanForm,
    ITKImportForm,
)

from .models import (
    WeeklyPlan,
    SAPOrder,
    SAPOperation,
    ITKRequest,
)

from capacity.models import WorkCenter


SAP_COLUMNS = [
    "Priority",
    "Oper. System Status",
    "Order",
    "User Status",
    "MaintActivityType",
    "Operation WorkCenter",
    "Operation/Activity",
    "Operation short text",
    "System status",
    "Number",
    "Actual work",
    "Work",
    "Revision",
    "Created on",
    "Order Type",
]


REQUIRED_COLUMNS = [
    "Order",
    "Operation/Activity",
    "Operation WorkCenter",
    "Number",
    "Work",
]


def clean_text(value):
    if value is None:
        return ""

    return str(value).strip()


def clean_order_number(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def clean_operation_number(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    text = str(value).strip()

    if text.isdigit():
        return text.zfill(4)

    return text


def clean_decimal(value):
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def clean_integer(value):
    if value is None or value == "":
        return None

    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def clean_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if hasattr(value, "year") and hasattr(value, "month"):
        return value

    text = str(value).strip()

    formats = [
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format
            ).date()
        except ValueError:
            continue

    return None


def read_sap_excel(excel_file):
    workbook = load_workbook(
        excel_file,
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active

    first_row = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1
        )
    )

    headers = []

    for cell in first_row:
        if isinstance(cell.value, str):
            headers.append(cell.value.strip())
        else:
            headers.append(cell.value)

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in headers
    ]

    if missing_columns:
        workbook.close()

        return {
            "success": False,
            "missing_columns": missing_columns,
        }

    column_map = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    known_workcenters = set(
        WorkCenter.objects.values_list(
            "code",
            flat=True
        )
    )

    total_rows = 0
    valid_rows = 0
    error_rows = 0

    unique_orders = set()
    unknown_workcenters = set()

    problems = []
    valid_data = []

    for excel_row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):
        if not any(
            value is not None
            for value in row
        ):
            continue

        total_rows += 1

        def get_value(column_name):
            if column_name not in column_map:
                return None

            index = column_map[column_name]

            if index >= len(row):
                return None

            return row[index]

        order_number = clean_order_number(
            get_value("Order")
        )

        operation_number = clean_operation_number(
            get_value("Operation/Activity")
        )

        workcenter_code = clean_text(
            get_value("Operation WorkCenter")
        )

        required_people = clean_decimal(
            get_value("Number")
        )

        planned_work = clean_decimal(
            get_value("Work")
        )

        actual_work = clean_decimal(
            get_value("Actual work")
        )

        priority = clean_integer(
            get_value("Priority")
        )

        created_on = clean_date(
            get_value("Created on")
        )

        row_problems = []

        if not order_number:
            row_problems.append(
                "Sifariş nömrəsi boşdur."
            )

        if not operation_number:
            row_problems.append(
                "Operation nömrəsi boşdur."
            )

        if not workcenter_code:
            row_problems.append(
                "Ekip kodu boşdur."
            )

        if required_people is None:
            row_problems.append(
                "Tələb olunan adam sayı düzgün deyil."
            )
        elif required_people <= 0:
            row_problems.append(
                "Tələb olunan adam sayı 0 və ya mənfidir."
            )

        if planned_work is None:
            row_problems.append(
                "Planlanan MH düzgün deyil."
            )
        elif planned_work <= 0:
            row_problems.append(
                "Planlanan MH 0 və ya mənfidir."
            )

        if order_number:
            unique_orders.add(order_number)

        if (
            workcenter_code
            and workcenter_code not in known_workcenters
        ):
            unknown_workcenters.add(
                workcenter_code
            )

        if row_problems:
            error_rows += 1

            if len(problems) < 200:
                problems.append(
                    {
                        "row": excel_row_number,
                        "order": order_number or "-",
                        "operation": (
                            operation_number or "-"
                        ),
                        "problems": row_problems,
                    }
                )

            continue

        valid_rows += 1

        valid_data.append(
            {
                "priority": priority,

                "operation_system_status":
                    clean_text(
                        get_value(
                            "Oper. System Status"
                        )
                    ),

                "order_number": order_number,

                "user_status":
                    clean_text(
                        get_value("User Status")
                    ),

                "maintenance_activity_type":
                    clean_text(
                        get_value(
                            "MaintActivityType"
                        )
                    ),

                "workcenter_code":
                    workcenter_code,

                "operation_number":
                    operation_number,

                "description":
                    clean_text(
                        get_value(
                            "Operation short text"
                        )
                    ),

                "system_status":
                    clean_text(
                        get_value("System status")
                    ),

                "required_people":
                    required_people,

                "actual_work_mh":
                    actual_work,

                "planned_work_mh":
                    planned_work,

                "revision":
                    clean_text(
                        get_value("Revision")
                    ),

                "created_on":
                    created_on,

                "order_type":
                    clean_text(
                        get_value("Order Type")
                    ),
            }
        )

    workbook.close()

    return {
        "success": True,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "error_rows": error_rows,
        "unique_order_count": len(unique_orders),
        "unknown_workcenters": sorted(
            unknown_workcenters
        ),
        "unknown_workcenter_count": len(
            unknown_workcenters
        ),
        "problems": problems,
        "valid_data": valid_data,
    }


def sap_import(request):
    form = SAPImportForm()

    context = {
        "form": form
    }

    if request.method == "POST":
        form = SAPImportForm(
            request.POST,
            request.FILES
        )

        context["form"] = form

        if form.is_valid():
            excel_file = form.cleaned_data[
                "excel_file"
            ]

            action = request.POST.get(
                "action",
                "validate"
            )

            try:
                result = read_sap_excel(
                    excel_file
                )

                if not result["success"]:
                    context.update(
                        {
                            "validation_failed": True,
                            "file_name":
                                excel_file.name,
                            "missing_columns":
                                result[
                                    "missing_columns"
                                ],
                        }
                    )

                    return render(
                        request,
                        "planning/sap_import.html",
                        context
                    )

                context.update(
                    {
                        "validation_complete": True,
                        "file_name":
                            excel_file.name,
                        "total_rows":
                            result["total_rows"],
                        "valid_rows":
                            result["valid_rows"],
                        "error_rows":
                            result["error_rows"],
                        "unique_order_count":
                            result[
                                "unique_order_count"
                            ],
                        "unknown_workcenters":
                            result[
                                "unknown_workcenters"
                            ],
                        "unknown_workcenter_count":
                            result[
                                "unknown_workcenter_count"
                            ],
                        "problems":
                            result["problems"],
                    }
                )

                if action == "import":
                    if result["error_rows"] > 0:
                        context[
                            "import_blocked"
                        ] = True

                        return render(
                            request,
                            "planning/sap_import.html",
                            context
                        )

                    imported_orders = 0
                    imported_operations = 0
                    updated_operations = 0

                    with transaction.atomic():
                        for item in result[
                            "valid_data"
                        ]:
                            order, order_created = (
                                SAPOrder.objects
                                .get_or_create(
                                    order_number=item[
                                        "order_number"
                                    ],
                                    defaults={
                                        "priority":
                                            item[
                                                "priority"
                                            ],
                                        "order_type":
                                            item[
                                                "order_type"
                                            ],
                                        "created_on":
                                            item[
                                                "created_on"
                                            ],
                                        "system_status":
                                            item[
                                                "system_status"
                                            ],
                                        "revision":
                                            item[
                                                "revision"
                                            ],
                                    }
                                )
                            )

                            if order_created:
                                imported_orders += 1
                            else:
                                order.priority = item[
                                    "priority"
                                ]
                                order.order_type = item[
                                    "order_type"
                                ]
                                order.created_on = item[
                                    "created_on"
                                ]
                                order.system_status = (
                                    item[
                                        "system_status"
                                    ]
                                )
                                order.revision = item[
                                    "revision"
                                ]
                                order.save()

                            workcenter = (
                                WorkCenter.objects
                                .filter(
                                    code=item[
                                        "workcenter_code"
                                    ]
                                )
                                .first()
                            )

                            operation, created = (
                                SAPOperation.objects
                                .update_or_create(
                                    order=order,
                                    operation_number=item[
                                        "operation_number"
                                    ],
                                    defaults={
                                        "work_center":
                                            workcenter,

                                        "work_center_code_raw":
                                            item[
                                                "workcenter_code"
                                            ],

                                        "description":
                                            item[
                                                "description"
                                            ],

                                        "required_people":
                                            item[
                                                "required_people"
                                            ],

                                        "planned_work_mh":
                                            item[
                                                "planned_work_mh"
                                            ],

                                        "actual_work_mh":
                                            item[
                                                "actual_work_mh"
                                            ],

                                        "operation_system_status":
                                            item[
                                                "operation_system_status"
                                            ],

                                        "user_status":
                                            item[
                                                "user_status"
                                            ],

                                        "maintenance_activity_type":
                                            item[
                                                "maintenance_activity_type"
                                            ],

                                        "revision":
                                            item[
                                                "revision"
                                            ],
                                    }
                                )
                            )

                            if created:
                                imported_operations += 1
                            else:
                                updated_operations += 1

                    context[
                        "import_success"
                    ] = True

                    context[
                        "imported_orders"
                    ] = imported_orders

                    context[
                        "imported_operations"
                    ] = imported_operations

                    context[
                        "updated_operations"
                    ] = updated_operations

            except Exception as error:
                context.update(
                    {
                        "validation_failed": True,
                        "general_error": str(error),
                    }
                )

    return render(
        request,
        "planning/sap_import.html",
        context
    )


def download_sap_template(request):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "SAP məlumatları"

    for column_number, column_name in enumerate(
        SAP_COLUMNS,
        start=1
    ):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=column_name
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    example_row = [
        2,
        "REL",
        "640000000001",
        "PLAN",
        "",
        "MEXANK-1",
        "0010",
        "Nasosun mexaniki təmiri",
        "REL",
        4,
        0,
        27,
        "W35-2026",
        "24.08.2026",
        "ZM01",
    ]

    for column_number, value in enumerate(
        example_row,
        start=1
    ):
        worksheet.cell(
            row=2,
            column=column_number,
            value=value
        )

    widths = {
        "A": 12,
        "B": 22,
        "C": 20,
        "D": 20,
        "E": 22,
        "F": 24,
        "G": 22,
        "H": 45,
        "I": 25,
        "J": 14,
        "K": 16,
        "L": 16,
        "M": 18,
        "N": 18,
        "O": 16,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:O2"

    instruction_sheet = (
        workbook.create_sheet(
            "Təlimat"
        )
    )

    instructions = [
        [
            "MPS — SAP / IW37N məlumat şablonu"
        ],
        [""],
        ["Sahə", "Açıqlama"],
        [
            "Order",
            "SAP təmir sifarişinin nömrəsi. Məcburidir."
        ],
        [
            "Operation/Activity",
            "Operation nömrəsi. Məcburidir."
        ],
        [
            "Operation WorkCenter",
            "İşi icra edən ekip. Məcburidir."
        ],
        [
            "Number",
            "Operation üçün tələb olunan adam sayı."
        ],
        [
            "Work",
            "Planlanan adam-saat (MH)."
        ],
        [
            "Priority",
            "SAP prioriteti."
        ],
        [
            "Created on",
            "Sifarişin yaranma tarixi."
        ],
        [
            "Revision",
            "SAP revision/həftə məlumatı."
        ],
        [
            "User Status",
            "PLAN, OPEG və digər statuslar."
        ],
        [
            "Qeyd",
            "Sütun adlarını dəyişməyin."
        ],
        [
            "Qeyd",
            "Hər sətir bir operation olmalıdır."
        ],
    ]

    for row in instructions:
        instruction_sheet.append(row)

    instruction_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    instruction_sheet["A3"].font = Font(
        bold=True
    )

    instruction_sheet["B3"].font = Font(
        bold=True
    )

    instruction_sheet.column_dimensions[
        "A"
    ].width = 28

    instruction_sheet.column_dimensions[
        "B"
    ].width = 70

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="MPS_SAP_IW37N_Sablonu.xlsx"'
    )

    return response
def weekly_plan_list(request):

    weekly_plans = WeeklyPlan.objects.all()

    return render(
        request,
        "planning/weekly_plan_list.html",
        {
            "weekly_plans": weekly_plans
        }
    )


def weekly_plan_create(request):

    if request.method == "POST":

        form = WeeklyPlanForm(
            request.POST
        )

        if form.is_valid():

            weekly_plan = form.save()

            return render(
                request,
                "planning/weekly_plan_created.html",
                {
                    "weekly_plan": weekly_plan
                }
            )

    else:

        form = WeeklyPlanForm()

    return render(
        request,
        "planning/weekly_plan_create.html",
        {
            "form": form
        }
    )
def clean_itk_order_number(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def itk_import(request):
    form = ITKImportForm()

    context = {
        "form": form
    }

    if request.method == "POST":
        form = ITKImportForm(
            request.POST,
            request.FILES
        )

        context["form"] = form

        if form.is_valid():
            weekly_plan = form.cleaned_data["weekly_plan"]
            excel_file = form.cleaned_data["excel_file"]

            action = request.POST.get(
                "action",
                "validate"
            )

            try:
                workbook = load_workbook(
                    excel_file,
                    read_only=True,
                    data_only=True
                )

                worksheet = workbook.active

                first_row = next(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=1
                    )
                )

                headers = [
                    str(cell.value).strip()
                    if cell.value is not None
                    else ""
                    for cell in first_row
                ]

                required_columns = [
                    "Order"
                ]

                missing_columns = [
                    column
                    for column in required_columns
                    if column not in headers
                ]

                if missing_columns:
                    workbook.close()

                    context.update({
                        "validation_failed": True,
                        "missing_columns": missing_columns,
                    })

                    return render(
                        request,
                        "planning/itk_import.html",
                        context
                    )

                column_map = {
                    header: index
                    for index, header in enumerate(headers)
                    if header
                }

                rows = []
                problems = []
                seen_orders = set()

                sap_orders = set(
                    SAPOrder.objects.values_list(
                        "order_number",
                        flat=True
                    )
                )

                found_in_sap = []
                missing_in_sap = []

                for excel_row, row in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        values_only=True
                    ),
                    start=2
                ):
                    if not any(
                        value is not None
                        for value in row
                    ):
                        continue

                    order_number = clean_itk_order_number(
                        row[column_map["Order"]]
                    )

                    note = ""

                    if "Qeyd" in column_map:
                        note_value = row[
                            column_map["Qeyd"]
                        ]

                        if note_value is not None:
                            note = str(note_value).strip()

                    if not order_number:
                        problems.append({
                            "row": excel_row,
                            "order": "-",
                            "reason": "Sifariş nömrəsi boşdur."
                        })
                        continue

                    if order_number in seen_orders:
                        problems.append({
                            "row": excel_row,
                            "order": order_number,
                            "reason": (
                                "Sifariş Excel-də təkrarlanır."
                            )
                        })
                        continue

                    seen_orders.add(order_number)

                    exists_in_sap = (
                        order_number in sap_orders
                    )

                    item = {
                        "row": excel_row,
                        "order_number": order_number,
                        "note": note,
                        "exists_in_sap": exists_in_sap,
                    }

                    rows.append(item)

                    if exists_in_sap:
                        found_in_sap.append(item)
                    else:
                        missing_in_sap.append(item)

                workbook.close()

                context.update({
                    "validation_complete": True,
                    "weekly_plan": weekly_plan,
                    "total_orders": len(rows),
                    "found_count": len(found_in_sap),
                    "missing_count": len(missing_in_sap),
                    "problem_count": len(problems),
                    "found_in_sap": found_in_sap,
                    "missing_in_sap": missing_in_sap,
                    "problems": problems,
                })

                if action == "import":
                    if problems:
                        context["import_blocked"] = True

                    else:
                        created_count = 0
                        updated_count = 0

                        with transaction.atomic():
                            for item in rows:
                                request_item, created = (
                                    ITKRequest.objects.update_or_create(
                                        weekly_plan=weekly_plan,
                                        order_number=item[
                                            "order_number"
                                        ],
                                        defaults={
                                            "note": item["note"]
                                        }
                                    )
                                )

                                if created:
                                    created_count += 1
                                else:
                                    updated_count += 1

                        context.update({
                            "import_success": True,
                            "created_count": created_count,
                            "updated_count": updated_count,
                        })

            except Exception as error:
                context.update({
                    "validation_failed": True,
                    "general_error": str(error),
                })

    return render(
        request,
        "planning/itk_import.html",
        context
    )


def download_itk_template(request):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "İTK sifarişləri"

    headers = [
        "Order",
        "Qeyd",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    worksheet.append([
        "640000000001",
        "İTK tərəfindən proqrama tələb olunur"
    ])

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 60

    worksheet.freeze_panes = "A2"

    instruction = workbook.create_sheet(
        "Təlimat"
    )

    instruction.append([
        "MPS — İTK Həftəlik Proqram Şablonu"
    ])

    instruction.append([])

    instruction.append([
        "Order",
        (
            "İTK tərəfindən həftəlik proqrama "
            "daxil edilməsi tələb olunan SAP "
            "sifariş nömrəsi."
        )
    ])

    instruction.append([
        "Qeyd",
        "İstəyə bağlı İTK qeydi."
    ])

    instruction.append([])

    instruction.append([
        "Vacib",
        (
            "Hər sifariş nömrəsini yalnız "
            "bir dəfə daxil edin."
        )
    ])

    instruction.append([
        "Vacib",
        (
            "MPS sifariş nömrəsini SAP "
            "məlumatları ilə avtomatik tutuşduracaq."
        )
    ])

    instruction.column_dimensions["A"].width = 25
    instruction.column_dimensions["B"].width = 75

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; '
        'filename="MPS_ITK_Sifaris_Sablonu.xlsx"'
    )

    return response

def itk_operations(request, plan_id):
    weekly_plan = WeeklyPlan.objects.filter(
        id=plan_id
    ).first()

    if weekly_plan is None:
        return HttpResponse(
            "Həftəlik proqram tapılmadı.",
            status=404
        )

    itk_requests = (
        ITKRequest.objects
        .filter(weekly_plan=weekly_plan)
        .order_by("order_number")
    )

    results = []

    total_itk_orders = 0
    found_orders = 0
    missing_orders = 0
    total_operations = 0
    total_planned_mh = Decimal("0")

    workcenter_summary = {}

    for itk_request in itk_requests:

        total_itk_orders += 1

        sap_order = (
            SAPOrder.objects
            .filter(
                order_number=itk_request.order_number
            )
            .first()
        )

        if sap_order is None:

            missing_orders += 1

            results.append({
                "itk_request": itk_request,
                "sap_order": None,
                "operations": [],
                "status": "NOT_FOUND",
                "status_text": (
                    "SAP-da tapılmadı — "
                    "planlamaya daxil edilə bilməz"
                ),
            })

            continue

        found_orders += 1

        operations = list(
            SAPOperation.objects
            .filter(order=sap_order)
            .select_related("work_center")
            .order_by("operation_number")
        )

        operation_rows = []

        for operation in operations:

            total_operations += 1

            planned_mh = (
                operation.planned_work_mh
                or Decimal("0")
            )

            total_planned_mh += planned_mh

            if operation.work_center:
                workcenter_code = (
                    operation.work_center.code
                )
            else:
                workcenter_code = (
                    operation.work_center_code_raw
                    or "Təyin edilməyib"
                )

            if workcenter_code not in workcenter_summary:
                workcenter_summary[
                    workcenter_code
                ] = {
                    "operation_count": 0,
                    "planned_mh": Decimal("0"),
                }

            workcenter_summary[
                workcenter_code
            ]["operation_count"] += 1

            workcenter_summary[
                workcenter_code
            ]["planned_mh"] += planned_mh

            duration = (
                operation.calculated_duration_hours
            )

            operation_rows.append({
                "operation": operation,
                "workcenter_code": workcenter_code,
                "duration": duration,
            })

        if operations:
            status = "FOUND"
            status_text = "SAP-da tapıldı"
        else:
            status = "NO_OPERATIONS"
            status_text = (
                "Sifariş SAP-da var, "
                "operation tapılmadı"
            )

        results.append({
            "itk_request": itk_request,
            "sap_order": sap_order,
            "operations": operation_rows,
            "status": status,
            "status_text": status_text,
        })

    workcenter_rows = []

    for code, data in sorted(
        workcenter_summary.items()
    ):
        workcenter_rows.append({
            "code": code,
            "operation_count": (
                data["operation_count"]
            ),
            "planned_mh": (
                data["planned_mh"]
            ),
        })

    context = {
        "weekly_plan": weekly_plan,
        "results": results,

        "total_itk_orders": total_itk_orders,
        "found_orders": found_orders,
        "missing_orders": missing_orders,
        "total_operations": total_operations,
        "total_planned_mh": total_planned_mh,

        "workcenter_rows": workcenter_rows,
    }

    return render(
        request,
        "planning/itk_operations.html",
        context
    )