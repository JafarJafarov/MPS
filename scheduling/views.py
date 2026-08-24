from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from datetime import timedelta

from decimal import Decimal

from django.contrib import messages
from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)

from planning.models import WeeklyPlan
from capacity.models import WeeklyCapacity

from .models import (
    ScheduledOperation,
    SchedulingRun,
)

from .scheduler import WeeklySchedulingEngine


# =========================================================
# SCHEDULING DASHBOARD
# =========================================================

def scheduling_dashboard(
    request,
    plan_id
):
    """
    Seçilmiş həftə üçün əsas Scheduling Dashboard.

    Buradan:
    - scheduler işə salınır,
    - planlanan operation-lar göstərilir,
    - planlanmayan operation-lar göstərilir,
    - manual review işləri göstərilir,
    - ekip capacity/utilization hesablanır,
    - Gantt segmentləri göstərilir.
    """

    weekly_plan = get_object_or_404(
        WeeklyPlan,
        id=plan_id
    )

    # =====================================================
    # SCHEDULER-İ İŞƏ SAL
    # =====================================================

    if request.method == "POST":

        try:

            engine = WeeklySchedulingEngine(
                weekly_plan
            )

            scheduling_run = engine.run()

            messages.success(
                request,
                (
                    "Scheduling uğurla tamamlandı. "
                    f"{scheduling_run.scheduled_operations} "
                    "operation proqrama alındı, "
                    f"{scheduling_run.not_scheduled_operations} "
                    "operation proqrama alınmadı, "
                    f"{scheduling_run.review_operations} "
                    "operation manual review tələb edir."
                )
            )

        except Exception as error:

            messages.error(
                request,
                (
                    "Scheduling zamanı xəta baş verdi: "
                    f"{error}"
                )
            )

        return redirect(
            "scheduling:dashboard",
            plan_id=weekly_plan.id
        )

    # =====================================================
    # PROQRAMA ALINAN OPERATION-LAR
    # =====================================================

    scheduled = (
        ScheduledOperation.objects
        .filter(
            weekly_plan=weekly_plan,
            decision="SCHEDULED"
        )
        .select_related(
            "sap_operation",
            "sap_operation__order",
            "work_center",
        )
        .prefetch_related(
            "segments"
        )
        .order_by(
            "sequence"
        )
    )

    # =====================================================
    # PROQRAMA ALINMAYAN OPERATION-LAR
    # =====================================================

    not_scheduled = (
        ScheduledOperation.objects
        .filter(
            weekly_plan=weekly_plan,
            decision="NOT_SCHEDULED"
        )
        .select_related(
            "sap_operation",
            "sap_operation__order",
            "work_center",
        )
        .order_by(
            "sap_operation__order__order_number",
            "sap_operation__operation_number",
        )
    )

    # =====================================================
    # MANUAL REVIEW
    # =====================================================

    review = (
        ScheduledOperation.objects
        .filter(
            weekly_plan=weekly_plan,
            decision="REVIEW"
        )
        .select_related(
            "sap_operation",
            "sap_operation__order",
            "work_center",
        )
        .order_by(
            "sap_operation__order__order_number",
            "sap_operation__operation_number",
        )
    )

    # =====================================================
    # SON SCHEDULING RUN
    # =====================================================

    last_run = (
        SchedulingRun.objects
        .filter(
            weekly_plan=weekly_plan
        )
        .order_by(
            "-started_at"
        )
        .first()
    )

    # =====================================================
    # ÜMUMİ PLANLANAN MH
    # =====================================================

    total_planned_mh = Decimal("0")

    for item in scheduled:

        if item.planned_mh is not None:

            total_planned_mh += (
                item.planned_mh
            )

    # =====================================================
    # WEEKLY CAPACITY
    # =====================================================

    weekly_capacities = (
        WeeklyCapacity.objects
        .filter(
            weekly_plan=weekly_plan
        )
        .select_related(
            "work_center"
        )
        .order_by(
            "work_center__code"
        )
    )

    # =====================================================
    # EKİP ÜZRƏ PLANLANAN MH
    # =====================================================

    planned_by_workcenter = {}

    for item in scheduled:

        if item.work_center is None:
            continue

        work_center_id = (
            item.work_center_id
        )

        if (
            work_center_id
            not in planned_by_workcenter
        ):

            planned_by_workcenter[
                work_center_id
            ] = {
                "planned_mh": Decimal("0"),
                "operations": 0,
            }

        planned_by_workcenter[
            work_center_id
        ]["planned_mh"] += (
            item.planned_mh
            or Decimal("0")
        )

        planned_by_workcenter[
            work_center_id
        ]["operations"] += 1

    # =====================================================
    # CAPACITY / UTILIZATION CƏDVƏLİ
    # =====================================================

    workcenter_rows = []

    total_capacity_mh = Decimal("0")

    for capacity in weekly_capacities:

        work_center = (
            capacity.work_center
        )

        weekly_capacity_mh = (
            Decimal(capacity.headcount)
            * capacity.productive_hours_per_day
            * Decimal(capacity.working_days)
        )

        total_capacity_mh += (
            weekly_capacity_mh
        )

        plan_data = (
            planned_by_workcenter.get(
                work_center.id,
                {
                    "planned_mh": Decimal("0"),
                    "operations": 0,
                }
            )
        )

        planned_mh = (
            plan_data["planned_mh"]
        )

        remaining_mh = (
            weekly_capacity_mh
            - planned_mh
        )

        if remaining_mh < 0:
            remaining_mh = Decimal("0")

        if weekly_capacity_mh > 0:

            utilization_percent = (
                planned_mh
                / weekly_capacity_mh
                * Decimal("100")
            )

        else:

            utilization_percent = Decimal("0")

        utilization_percent = (
            utilization_percent.quantize(
                Decimal("0.01")
            )
        )

        workcenter_rows.append(
            {
                "work_center": work_center,

                "code": work_center.code,

                "headcount": (
                    capacity.headcount
                ),

                "productive_hours_per_day": (
                    capacity
                    .productive_hours_per_day
                ),

                "working_days": (
                    capacity.working_days
                ),

                "weekly_capacity_mh": (
                    weekly_capacity_mh
                ),

                "planned_mh": (
                    planned_mh
                ),

                "remaining_mh": (
                    remaining_mh
                ),

                "utilization_percent": (
                    utilization_percent
                ),

                "operations": (
                    plan_data["operations"]
                ),
            }
        )

    # =====================================================
    # ÜMUMİ UTILIZATION
    # =====================================================

    if total_capacity_mh > 0:

        total_utilization = (
            total_planned_mh
            / total_capacity_mh
            * Decimal("100")
        )

        total_utilization = (
            total_utilization.quantize(
                Decimal("0.01")
            )
        )

    else:

        total_utilization = Decimal("0")

    # =====================================================
    # GECİKMİŞ İŞLƏR
    # =====================================================

    overdue_count = (
        scheduled.filter(
            is_overdue=True
        ).count()
    )

    # =====================================================
    # 11.3 REAL GANTT TIMELINE
    # =====================================================
    # Timeline: 08:00-17:00. Lunch (12:00-13:00) remains
    # visible as a separate zone. Segment positions are
    # calculated server-side so the template stays simple.

    gantt_day_start_minute = 8 * 60
    gantt_day_end_minute = 17 * 60
    gantt_day_minutes = (
        gantt_day_end_minute
        - gantt_day_start_minute
    )

    gantt_days = []
    current_date = weekly_plan.start_date

    while current_date <= weekly_plan.end_date:
        gantt_days.append(
            {
                "date": current_date,
                "label": current_date.strftime("%d.%m"),
                "weekday": current_date.strftime("%a"),
            }
        )
        current_date += timedelta(days=1)

    gantt_rows = []

    for item in scheduled:
        segment_items = []

        for segment in item.segments.all():

            start_minute = (
                segment.start_time.hour * 60
                + segment.start_time.minute
            )
            end_minute = (
                segment.end_time.hour * 60
                + segment.end_time.minute
            )

            start_minute = max(
                start_minute,
                gantt_day_start_minute,
            )
            end_minute = min(
                end_minute,
                gantt_day_end_minute,
            )

            if end_minute <= start_minute:
                continue

            left_percent = (
                Decimal(
                    start_minute
                    - gantt_day_start_minute
                )
                / Decimal(gantt_day_minutes)
                * Decimal("100")
            )

            width_percent = (
                Decimal(
                    end_minute
                    - start_minute
                )
                / Decimal(gantt_day_minutes)
                * Decimal("100")
            )

            segment_items.append(
                {
                    "date": segment.work_date,
                    "start_time": segment.start_time,
                    "end_time": segment.end_time,
                    "left_percent": left_percent.quantize(
                        Decimal("0.01")
                    ),
                    "width_percent": width_percent.quantize(
                        Decimal("0.01")
                    ),
                }
            )

        gantt_rows.append(
            {
                "item": item,
                "segments": segment_items,
            }
        )

    # =====================================================
    # 11.5 MANAGEMENT KPI SUMMARY
    # =====================================================

    total_decisions = (
        scheduled.count()
        + not_scheduled.count()
        + review.count()
    )

    if total_decisions > 0:
        schedule_success_percent = (
            Decimal(scheduled.count())
            / Decimal(total_decisions)
            * Decimal("100")
        ).quantize(Decimal("0.01"))

        overdue_percent = (
            Decimal(overdue_count)
            / Decimal(total_decisions)
            * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        schedule_success_percent = Decimal("0")
        overdue_percent = Decimal("0")

    remaining_capacity_mh = (
        total_capacity_mh
        - total_planned_mh
    )

    overload_workcenters = sum(
        1
        for row in workcenter_rows
        if row["utilization_percent"] > Decimal("100")
    )

    high_load_workcenters = sum(
        1
        for row in workcenter_rows
        if (
            row["utilization_percent"] >= Decimal("85")
            and row["utilization_percent"] <= Decimal("100")
        )
    )

    # Management status:
    # CRITICAL  -> overload və ya aşağı scheduling nəticəsi
    # ATTENTION -> unscheduled/review/overdue və ya yüksək ekip yükü
    # NORMAL    -> əsas risk görünmür
    if (
        overload_workcenters > 0
        or schedule_success_percent < Decimal("70")
    ):
        management_status = "CRITICAL"
        management_status_label = "KRİTİK"
    elif (
        not_scheduled.count() > 0
        or review.count() > 0
        or overdue_count > 0
        or high_load_workcenters > 0
        or total_utilization >= Decimal("85")
    ):
        management_status = "ATTENTION"
        management_status_label = "DİQQƏT"
    else:
        management_status = "NORMAL"
        management_status_label = "NORMAL"

    # =====================================================
    # DASHBOARD CONTEXT
    # =====================================================

    context = {

        "weekly_plan": (
            weekly_plan
        ),

        # ---------------------------------
        # Operation siyahıları
        # ---------------------------------

        "scheduled": (
            scheduled
        ),

        "not_scheduled": (
            not_scheduled
        ),

        "review": (
            review
        ),

        # ---------------------------------
        # Saylar
        # ---------------------------------

        "scheduled_count": (
            scheduled.count()
        ),

        "not_scheduled_count": (
            not_scheduled.count()
        ),

        "review_count": (
            review.count()
        ),

        "overdue_count": (
            overdue_count
        ),

        # ---------------------------------
        # MH / Capacity
        # ---------------------------------

        "total_planned_mh": (
            total_planned_mh
        ),

        "total_capacity_mh": (
            total_capacity_mh
        ),

        "total_utilization": (
            total_utilization
        ),

        "workcenter_rows": (
            workcenter_rows
        ),

        # ---------------------------------
        # 11.3 Gantt
        # ---------------------------------

        "gantt_days": gantt_days,
        "gantt_rows": gantt_rows,

        # ---------------------------------
        # 11.5 Management KPI
        # ---------------------------------

        "total_decisions": total_decisions,
        "schedule_success_percent": schedule_success_percent,
        "overdue_percent": overdue_percent,
        "remaining_capacity_mh": remaining_capacity_mh,
        "overload_workcenters": overload_workcenters,
        "high_load_workcenters": high_load_workcenters,
        "management_status": management_status,
        "management_status_label": management_status_label,

        # ---------------------------------
        # Run məlumatı
        # ---------------------------------

        "last_run": (
            last_run
        ),
    }

    return render(
        request,
        "scheduling/dashboard.html",
        context
    )

# =========================================================
# 11.6.1 EXCEL EXPORT
# =========================================================

def export_weekly_plan_excel(request, plan_id):
    """
    Seçilmiş həftənin MPS nəticələrini Excel faylına çıxarır.
    """

    weekly_plan = get_object_or_404(
        WeeklyPlan,
        id=plan_id,
    )

    results = (
        ScheduledOperation.objects
        .filter(weekly_plan=weekly_plan)
        .select_related(
            "sap_operation",
            "sap_operation__order",
            "work_center",
        )
        .prefetch_related("segments")
    )

    scheduled = results.filter(
        decision="SCHEDULED"
    ).order_by("sequence")

    exceptions = results.exclude(
        decision="SCHEDULED"
    ).order_by(
        "decision",
        "sap_operation__order__order_number",
        "sap_operation__operation_number",
    )

    capacities = (
        WeeklyCapacity.objects
        .filter(weekly_plan=weekly_plan)
        .select_related("work_center")
        .order_by("work_center__code")
    )

    workbook = Workbook()

    # -----------------------------------------------------
    # MANAGEMENT SUMMARY
    # -----------------------------------------------------

    ws = workbook.active
    ws.title = "Management Summary"

    ws.append([
        "MPS WEEKLY MANAGEMENT SUMMARY"
    ])

    ws.append([])
    ws.append(["Həftə", str(weekly_plan)])
    ws.append([
        "Tarix aralığı",
        f"{weekly_plan.start_date:%d.%m.%Y} - "
        f"{weekly_plan.end_date:%d.%m.%Y}",
    ])

    scheduled_count = scheduled.count()
    not_scheduled_count = results.filter(
        decision="NOT_SCHEDULED"
    ).count()
    review_count = results.filter(
        decision="REVIEW"
    ).count()

    total_planned_mh = sum(
        (
            item.planned_mh
            or Decimal("0")
        )
        for item in scheduled
    )

    total_capacity_mh = sum(
        (
            Decimal(capacity.headcount)
            * capacity.productive_hours_per_day
            * Decimal(capacity.working_days)
        )
        for capacity in capacities
    )

    total_decisions = results.count()

    if total_decisions:
        success_percent = (
            Decimal(scheduled_count)
            / Decimal(total_decisions)
            * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        success_percent = Decimal("0")

    if total_capacity_mh:
        utilization = (
            total_planned_mh
            / total_capacity_mh
            * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        utilization = Decimal("0")

    ws.append([])
    ws.append(["KPI", "Dəyər"])
    ws.append(["Scheduled Operations", scheduled_count])
    ws.append(["Not Scheduled", not_scheduled_count])
    ws.append(["Manual Review", review_count])
    ws.append(["Schedule Success %", float(success_percent)])
    ws.append(["Total Planned MH", float(total_planned_mh)])
    ws.append(["Total Capacity MH", float(total_capacity_mh)])
    ws.append(["Capacity Utilization %", float(utilization)])
    ws.append([
        "Remaining Capacity MH",
        float(total_capacity_mh - total_planned_mh),
    ])
    ws.append([
        "Overdue Operations",
        scheduled.filter(is_overdue=True).count(),
    ])

    # -----------------------------------------------------
    # SCHEDULED OPERATIONS
    # -----------------------------------------------------

    ws_scheduled = workbook.create_sheet(
        "Scheduled Operations"
    )

    ws_scheduled.append([
        "Order",
        "Operation",
        "Description",
        "WorkCenter",
        "Priority",
        "Due Date",
        "Overdue",
        "Required People",
        "Planned MH",
        "Duration Hours",
        "Scheduled Date",
        "Start",
        "End",
        "Sequence",
        "Locked",
    ])

    for item in scheduled:
        operation = item.sap_operation

        ws_scheduled.append([
            operation.order.order_number,
            operation.operation_number,
            operation.description,
            (
                item.work_center.code
                if item.work_center
                else ""
            ),
            item.priority,
            item.due_date,
            "YES" if item.is_overdue else "NO",
            float(item.required_people)
            if item.required_people is not None
            else None,
            float(item.planned_mh)
            if item.planned_mh is not None
            else None,
            float(item.planned_duration_hours)
            if item.planned_duration_hours is not None
            else None,
            item.scheduled_date,
            item.scheduled_start,
            item.scheduled_end,
            item.sequence,
            "YES" if item.is_locked else "NO",
        ])

    # -----------------------------------------------------
    # CAPACITY
    # -----------------------------------------------------

    ws_capacity = workbook.create_sheet("Capacity")

    ws_capacity.append([
        "WorkCenter",
        "Headcount",
        "Productive Hours / Day",
        "Working Days",
        "Capacity MH",
        "Planned MH",
        "Remaining MH",
        "Utilization %",
    ])

    for capacity in capacities:
        capacity_mh = (
            Decimal(capacity.headcount)
            * capacity.productive_hours_per_day
            * Decimal(capacity.working_days)
        )

        planned_mh = sum(
            (
                item.planned_mh
                or Decimal("0")
            )
            for item in scheduled
            if item.work_center_id
            == capacity.work_center_id
        )

        remaining_mh = (
            capacity_mh
            - planned_mh
        )

        utilization_percent = (
            planned_mh
            / capacity_mh
            * Decimal("100")
            if capacity_mh
            else Decimal("0")
        )

        ws_capacity.append([
            capacity.work_center.code,
            capacity.headcount,
            float(capacity.productive_hours_per_day),
            capacity.working_days,
            float(capacity_mh),
            float(planned_mh),
            float(remaining_mh),
            float(
                utilization_percent.quantize(
                    Decimal("0.01")
                )
            ),
        ])

    # -----------------------------------------------------
    # EXCEPTIONS
    # -----------------------------------------------------

    ws_exceptions = workbook.create_sheet(
        "Exceptions"
    )

    ws_exceptions.append([
        "Decision",
        "Order",
        "Operation",
        "Description",
        "WorkCenter",
        "Priority",
        "Due Date",
        "Required People",
        "Planned MH",
        "Reason",
    ])

    for item in exceptions:
        operation = item.sap_operation

        ws_exceptions.append([
            item.decision,
            operation.order.order_number,
            operation.operation_number,
            operation.description,
            (
                item.work_center.code
                if item.work_center
                else ""
            ),
            item.priority,
            item.due_date,
            float(item.required_people)
            if item.required_people is not None
            else None,
            float(item.planned_mh)
            if item.planned_mh is not None
            else None,
            item.decision_reason,
        ])

    # -----------------------------------------------------
    # GANTT SEGMENTS
    # -----------------------------------------------------

    ws_gantt = workbook.create_sheet(
        "Gantt Segments"
    )

    ws_gantt.append([
        "Order",
        "Operation",
        "WorkCenter",
        "Date",
        "Start",
        "End",
        "Required People",
        "Segment MH",
        "Segment Sequence",
    ])

    for item in scheduled:
        for segment in item.segments.all():
            ws_gantt.append([
                item.sap_operation.order.order_number,
                item.sap_operation.operation_number,
                (
                    item.work_center.code
                    if item.work_center
                    else ""
                ),
                segment.work_date,
                segment.start_time,
                segment.end_time,
                float(segment.required_people),
                float(segment.segment_mh),
                segment.sequence,
            ])

    # -----------------------------------------------------
    # VISUAL GANTT CHART
    # -----------------------------------------------------

    ws_visual = workbook.create_sheet("Gantt Chart")

    gantt_start_hour = 8
    gantt_end_hour = 17
    slot_minutes = 15

    # Build the complete weekly timeline.
    timeline = []
    current_date = weekly_plan.start_date

    while current_date <= weekly_plan.end_date:
        if current_date.weekday() < 5:
            current_time_minutes = gantt_start_hour * 60

            while current_time_minutes < gantt_end_hour * 60:
                hour = current_time_minutes // 60
                minute = current_time_minutes % 60

                timeline.append(
                    (
                        current_date,
                        hour,
                        minute,
                    )
                )

                current_time_minutes += slot_minutes

        current_date += timedelta(days=1)

    fixed_headers = [
        "Order",
        "Operation",
        "Description",
        "Ekip",
        "Priority",
        "Planned MH",
    ]

    for column_index, header in enumerate(
        fixed_headers,
        start=1,
    ):
        ws_visual.cell(
            row=1,
            column=column_index,
            value=header,
        )
        ws_visual.merge_cells(
            start_row=1,
            start_column=column_index,
            end_row=2,
            end_column=column_index,
        )

    timeline_start_col = len(fixed_headers) + 1

    # Day headers and time headers.
    timeline_col = timeline_start_col
    timeline_by_key = {}

    current_date = None
    day_start_col = None

    for date_value, hour, minute in timeline:

        if date_value != current_date:

            if current_date is not None:
                ws_visual.merge_cells(
                    start_row=1,
                    start_column=day_start_col,
                    end_row=1,
                    end_column=timeline_col - 1,
                )

            current_date = date_value
            day_start_col = timeline_col

            ws_visual.cell(
                row=1,
                column=timeline_col,
                value=date_value,
            )

        ws_visual.cell(
            row=2,
            column=timeline_col,
            value=f"{hour:02d}:{minute:02d}",
        )

        timeline_by_key[
            (date_value, hour, minute)
        ] = timeline_col

        timeline_col += 1

    if current_date is not None:
        ws_visual.merge_cells(
            start_row=1,
            start_column=day_start_col,
            end_row=1,
            end_column=timeline_col - 1,
        )

    # Visual styles.
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    time_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    work_fill = PatternFill(
        fill_type="solid",
        fgColor="5B9BD5",
    )

    lunch_fill = PatternFill(
        fill_type="solid",
        fgColor="E7E6E6",
    )

    overdue_fill = PatternFill(
        fill_type="solid",
        fgColor="C00000",
    )

    locked_fill = PatternFill(
        fill_type="solid",
        fgColor="70AD47",
    )

    thin_side = Side(
        style="thin",
        color="D9E1F2",
    )

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row_number in (1, 2):
        for cell in ws_visual[row_number]:
            cell.fill = (
                header_fill
                if row_number == 1
                else time_fill
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF"
                if row_number == 1
                else "000000",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                text_rotation=0,
            )
            cell.border = thin_border

    # Fill lunch interval across every workday.
    for (
        date_value,
        hour,
        minute,
    ) in timeline:
        column = timeline_by_key[
            (date_value, hour, minute)
        ]

        if (
            hour == 12
        ):
            ws_visual.cell(
                row=2,
                column=column,
            ).fill = lunch_fill

    # Operation rows.
    gantt_row = 3

    for item in scheduled:

        operation = item.sap_operation

        ws_visual.cell(
            row=gantt_row,
            column=1,
            value=operation.order.order_number,
        )
        ws_visual.cell(
            row=gantt_row,
            column=2,
            value=operation.operation_number,
        )
        ws_visual.cell(
            row=gantt_row,
            column=3,
            value=operation.description or "",
        )
        ws_visual.cell(
            row=gantt_row,
            column=4,
            value=(
                item.work_center.code
                if item.work_center
                else ""
            ),
        )
        ws_visual.cell(
            row=gantt_row,
            column=5,
            value=(
                f"P{item.priority}"
                if item.priority is not None
                else ""
            ),
        )
        ws_visual.cell(
            row=gantt_row,
            column=6,
            value=(
                float(item.planned_mh)
                if item.planned_mh is not None
                else None
            ),
        )

        for column in range(
            1,
            timeline_start_col,
        ):
            cell = ws_visual.cell(
                row=gantt_row,
                column=column,
            )
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        # Paint the lunch period.
        for (
            date_value,
            hour,
            minute,
        ) in timeline:

            column = timeline_by_key[
                (date_value, hour, minute)
            ]

            cell = ws_visual.cell(
                row=gantt_row,
                column=column,
            )

            cell.border = thin_border

            if hour == 12:
                cell.fill = lunch_fill

        # Paint actual schedule segments.
        for segment in item.segments.all():

            start_minutes = (
                segment.start_time.hour * 60
                + segment.start_time.minute
            )

            end_minutes = (
                segment.end_time.hour * 60
                + segment.end_time.minute
            )

            current_minutes = start_minutes

            while current_minutes < end_minutes:

                hour = current_minutes // 60
                minute = current_minutes % 60

                key = (
                    segment.work_date,
                    hour,
                    minute,
                )

                column = timeline_by_key.get(key)

                if column is not None:

                    cell = ws_visual.cell(
                        row=gantt_row,
                        column=column,
                    )

                    if item.is_overdue:
                        cell.fill = overdue_fill
                    elif item.is_locked:
                        cell.fill = locked_fill
                    else:
                        cell.fill = work_fill

                    cell.border = thin_border

                current_minutes += slot_minutes

        gantt_row += 1

    # Excel layout.
    ws_visual.freeze_panes = "G3"
    ws_visual.auto_filter.ref = (
        f"A2:{get_column_letter(timeline_col - 1)}"
        f"{max(gantt_row - 1, 2)}"
    )

    ws_visual.row_dimensions[1].height = 24
    ws_visual.row_dimensions[2].height = 30

    fixed_widths = {
        "A": 16,
        "B": 12,
        "C": 42,
        "D": 16,
        "E": 10,
        "F": 14,
    }

    for column_letter, width in fixed_widths.items():
        ws_visual.column_dimensions[
            column_letter
        ].width = width

    for column in range(
        timeline_start_col,
        timeline_col,
    ):
        ws_visual.column_dimensions[
            get_column_letter(column)
        ].width = 5

    # -----------------------------------------------------
    # BASIC EXCEL FORMATTING
    # -----------------------------------------------------

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                vertical="center"
            )

        for column_cells in worksheet.columns:

            max_length = 0
            column_number = (
                column_cells[0].column
            )

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(value)),
                )

            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = min(
                max(max_length + 2, 12),
                45,
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = (
        f"MPS_{weekly_plan.year}_"
        f"W{weekly_plan.week_number}.xlsx"
    )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}"'
    )

    return response


# =========================================================
# 11.6.2 PDF EXPORT
# =========================================================

def export_weekly_plan_pdf(request, plan_id):
    """
    Rəhbərlik üçün həftəlik MPS PDF hesabatı.
    """

    weekly_plan = get_object_or_404(
        WeeklyPlan,
        id=plan_id,
    )

    results = (
        ScheduledOperation.objects
        .filter(weekly_plan=weekly_plan)
        .select_related(
            "sap_operation",
            "sap_operation__order",
            "work_center",
        )
        .prefetch_related("segments")
    )

    scheduled = results.filter(
        decision="SCHEDULED"
    ).order_by("sequence")

    exceptions = results.exclude(
        decision="SCHEDULED"
    ).order_by(
        "decision",
        "sap_operation__order__order_number",
        "sap_operation__operation_number",
    )

    capacities = (
        WeeklyCapacity.objects
        .filter(weekly_plan=weekly_plan)
        .select_related("work_center")
        .order_by("work_center__code")
    )

    scheduled_count = scheduled.count()
    not_scheduled_count = results.filter(
        decision="NOT_SCHEDULED"
    ).count()
    review_count = results.filter(
        decision="REVIEW"
    ).count()
    total_decisions = results.count()

    total_planned_mh = sum(
        (
            item.planned_mh
            or Decimal("0")
        )
        for item in scheduled
    )

    total_capacity_mh = sum(
        (
            Decimal(capacity.headcount)
            * capacity.productive_hours_per_day
            * Decimal(capacity.working_days)
        )
        for capacity in capacities
    )

    if total_decisions:
        success_percent = (
            Decimal(scheduled_count)
            / Decimal(total_decisions)
            * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        success_percent = Decimal("0")

    if total_capacity_mh:
        utilization = (
            total_planned_mh
            / total_capacity_mh
            * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        utilization = Decimal("0")

    remaining_capacity = (
        total_capacity_mh
        - total_planned_mh
    )

    overdue_count = scheduled.filter(
        is_overdue=True
    ).count()

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"MPS {weekly_plan}",
        author="MPS",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "MPS_Title",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=17,
        leading=21,
        spaceAfter=8,
    )

    heading_style = ParagraphStyle(
        "MPS_Heading",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        spaceBefore=6,
        spaceAfter=6,
    )

    small_style = ParagraphStyle(
        "MPS_Small",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
    )

    story = []

    story.append(
        Paragraph(
            "MPS WEEKLY MAINTENANCE PROGRAM",
            title_style,
        )
    )

    story.append(
        Paragraph(
            (
                f"Həftə: {weekly_plan} &nbsp;&nbsp;&nbsp; "
                f"Tarix: {weekly_plan.start_date:%d.%m.%Y} - "
                f"{weekly_plan.end_date:%d.%m.%Y}"
            ),
            styles["BodyText"],
        )
    )

    story.append(Spacer(1, 5 * mm))

    # -----------------------------------------------------
    # MANAGEMENT SUMMARY
    # -----------------------------------------------------

    story.append(
        Paragraph(
            "Management Summary",
            heading_style,
        )
    )

    summary_data = [
        [
            "Scheduled",
            "Not Scheduled",
            "Review",
            "Success %",
            "Planned MH",
            "Capacity MH",
            "Utilization %",
            "Remaining MH",
            "Overdue",
        ],
        [
            str(scheduled_count),
            str(not_scheduled_count),
            str(review_count),
            f"{success_percent}%",
            f"{total_planned_mh}",
            f"{total_capacity_mh}",
            f"{utilization}%",
            f"{remaining_capacity}",
            str(overdue_count),
        ],
    ]

    summary_table = Table(
        summary_data,
        repeatRows=1,
    )

    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#344054")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D5DD")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    story.append(summary_table)
    story.append(Spacer(1, 5 * mm))

    # -----------------------------------------------------
    # CAPACITY
    # -----------------------------------------------------

    story.append(
        Paragraph(
            "Ekip Capacity / Utilization",
            heading_style,
        )
    )

    capacity_data = [
        [
            "WorkCenter",
            "Headcount",
            "Hours/Day",
            "Days",
            "Capacity MH",
            "Planned MH",
            "Remaining MH",
            "Utilization %",
        ]
    ]

    for capacity in capacities:
        capacity_mh = (
            Decimal(capacity.headcount)
            * capacity.productive_hours_per_day
            * Decimal(capacity.working_days)
        )

        planned_mh = sum(
            (
                item.planned_mh
                or Decimal("0")
            )
            for item in scheduled
            if item.work_center_id
            == capacity.work_center_id
        )

        remaining_mh = (
            capacity_mh
            - planned_mh
        )

        utilization_percent = (
            planned_mh
            / capacity_mh
            * Decimal("100")
            if capacity_mh
            else Decimal("0")
        ).quantize(Decimal("0.01"))

        capacity_data.append(
            [
                capacity.work_center.code,
                str(capacity.headcount),
                str(capacity.productive_hours_per_day),
                str(capacity.working_days),
                str(capacity_mh),
                str(planned_mh),
                str(remaining_mh),
                f"{utilization_percent}%",
            ]
        )

    capacity_table = Table(
        capacity_data,
        repeatRows=1,
    )

    capacity_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D0D5DD")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(capacity_table)
    story.append(PageBreak())

    # -----------------------------------------------------
    # SCHEDULED OPERATIONS
    # -----------------------------------------------------

    story.append(
        Paragraph(
            "Scheduled Operations",
            heading_style,
        )
    )

    scheduled_data = [
        [
            "Order",
            "Op.",
            "Description",
            "WC",
            "P",
            "Due Date",
            "People",
            "MH",
            "Date",
            "Start",
            "End",
        ]
    ]

    for item in scheduled:
        operation = item.sap_operation

        scheduled_data.append(
            [
                operation.order.order_number,
                operation.operation_number,
                Paragraph(
                    operation.description or "",
                    small_style,
                ),
                (
                    item.work_center.code
                    if item.work_center
                    else ""
                ),
                str(item.priority or ""),
                (
                    item.due_date.strftime("%d.%m.%Y")
                    if item.due_date
                    else ""
                ),
                str(item.required_people or ""),
                str(item.planned_mh or ""),
                (
                    item.scheduled_date.strftime("%d.%m.%Y")
                    if item.scheduled_date
                    else ""
                ),
                (
                    item.scheduled_start.strftime("%H:%M")
                    if item.scheduled_start
                    else ""
                ),
                (
                    item.scheduled_end.strftime("%H:%M")
                    if item.scheduled_end
                    else ""
                ),
            ]
        )

    scheduled_table = Table(
        scheduled_data,
        repeatRows=1,
        colWidths=[
            24 * mm,
            13 * mm,
            70 * mm,
            22 * mm,
            9 * mm,
            20 * mm,
            14 * mm,
            16 * mm,
            20 * mm,
            15 * mm,
            15 * mm,
        ],
    )

    scheduled_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D5DD")),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    story.append(scheduled_table)

    # -----------------------------------------------------
    # EXCEPTIONS
    # -----------------------------------------------------

    if exceptions.exists():
        story.append(PageBreak())

        story.append(
            Paragraph(
                "Exceptions / Manual Review",
                heading_style,
            )
        )

        exception_data = [
            [
                "Decision",
                "Order",
                "Op.",
                "WC",
                "P",
                "Due Date",
                "MH",
                "Reason",
            ]
        ]

        for item in exceptions:
            exception_data.append(
                [
                    item.decision,
                    item.sap_operation.order.order_number,
                    item.sap_operation.operation_number,
                    (
                        item.work_center.code
                        if item.work_center
                        else ""
                    ),
                    str(item.priority or ""),
                    (
                        item.due_date.strftime("%d.%m.%Y")
                        if item.due_date
                        else ""
                    ),
                    str(item.planned_mh or ""),
                    Paragraph(
                        item.decision_reason
                        or "Səbəb qeyd edilməyib.",
                        small_style,
                    ),
                ]
            )

        exception_table = Table(
            exception_data,
            repeatRows=1,
            colWidths=[
                25 * mm,
                27 * mm,
                15 * mm,
                22 * mm,
                10 * mm,
                22 * mm,
                16 * mm,
                120 * mm,
            ],
        )

        exception_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FEF3F2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D5DD")),
                    ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

        story.append(exception_table)

    document.build(story)

    pdf_data = buffer.getvalue()
    buffer.close()

    filename = (
        f"MPS_{weekly_plan.year}_"
        f"W{weekly_plan.week_number}.pdf"
    )

    response = HttpResponse(
        pdf_data,
        content_type="application/pdf",
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}"'
    )

    return response
