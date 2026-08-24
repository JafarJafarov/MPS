from io import BytesIO
from decimal import Decimal

from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .forms import CapacityImportForm
from .models import WorkCenter, WeeklyCapacity
from planning.models import WeeklyPlan


def capacity_import(request):
    form = CapacityImportForm()

    context = {
        "form": form,
        "weekly_plans": WeeklyPlan.objects.order_by("-year", "-week_number"),
    }

    if request.method == "POST":
        form = CapacityImportForm(
            request.POST,
            request.FILES
        )

        context["form"] = form

        if form.is_valid():

            weekly_plan_id = request.POST.get("weekly_plan")

            if not weekly_plan_id:
                context["validation_failed"] = True
                context["general_error"] = (
                    "Həftəlik proqram seçilməyib."
                )
                return render(
                    request,
                    "capacity/capacity_import.html",
                    context,
                )

            try:
                weekly_plan = WeeklyPlan.objects.get(
                    pk=weekly_plan_id
                )
            except (WeeklyPlan.DoesNotExist, ValueError, TypeError):
                context["validation_failed"] = True
                context["general_error"] = (
                    "Seçilmiş həftəlik proqram tapılmadı."
                )
                return render(
                    request,
                    "capacity/capacity_import.html",
                    context,
                )

            excel_file = form.cleaned_data[
                "excel_file"
            ]

            try:
                workbook = load_workbook(
                    excel_file,
                    read_only=True,
                    data_only=True
                )

                worksheet = workbook.active

                headers = [
                    str(cell.value).strip()
                    if cell.value is not None
                    else ""
                    for cell in next(
                        worksheet.iter_rows(
                            min_row=1,
                            max_row=1
                        )
                    )
                ]

                required_columns = [
                    "WorkCenter",
                    "Adam sayı",
                ]

                missing_columns = [
                    column
                    for column in required_columns
                    if column not in headers
                ]

                if missing_columns:
                    workbook.close()

                    context.update({
                        "validation_failed": True,
                        "missing_columns": missing_columns,
                    })

                    return render(
                        request,
                        "capacity/capacity_import.html",
                        context
                    )

                column_map = {
                    header: index
                    for index, header
                    in enumerate(headers)
                }

                rows = []
                problems = []
                seen_workcenters = set()

                for excel_row, row in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        values_only=True
                    ),
                    start=2
                ):

                    if not any(
                        value is not None
                        for value in row
                    ):
                        continue

                    code = row[
                        column_map["WorkCenter"]
                    ]

                    headcount = row[
                        column_map["Adam sayı"]
                    ]

                    code = (
                        str(code).strip().upper()
                        if code is not None
                        else ""
                    )

                    try:
                        headcount = int(
                            float(headcount)
                        )
                    except (TypeError, ValueError):
                        headcount = None

                    if not code:
                        problems.append(
                            f"Sətir {excel_row}: "
                            "WorkCenter boşdur."
                        )
                        continue

                    if code in seen_workcenters:
                        problems.append(
                            f"Sətir {excel_row}: "
                            f"{code} Excel-də təkrarlanır."
                        )
                        continue

                    seen_workcenters.add(code)

                    if (
                        headcount is None
                        or headcount <= 0
                    ):
                        problems.append(
                            f"Sətir {excel_row}: "
                            f"{code} üçün adam sayı "
                            "düzgün deyil."
                        )
                        continue

                    rows.append({
                        "code": code,
                        "headcount": headcount,
                    })

                workbook.close()

                context.update({
                    "validation_complete": True,
                    "weekly_plan": weekly_plan,
                    "total_rows":
                        len(rows) + len(problems),
                    "valid_rows": len(rows),
                    "error_rows": len(problems),
                    "problems": problems,
                })

                action = request.POST.get(
                    "action",
                    "validate"
                )

                if action == "import":

                    if problems:
                        context[
                            "import_blocked"
                        ] = True

                    else:
                        created_workcenters = 0
                        created_capacities = 0
                        updated_capacities = 0

                        with transaction.atomic():

                            for item in rows:

                                workcenter, wc_created = (
                                    WorkCenter.objects
                                    .get_or_create(
                                        code=item["code"],
                                        defaults={
                                            "is_active": True
                                        }
                                    )
                                )

                                if wc_created:
                                    created_workcenters += 1

                                capacity, cap_created = (
                                    WeeklyCapacity.objects
                                    .update_or_create(
                                        weekly_plan=
                                            weekly_plan,

                                        work_center=
                                            workcenter,

                                        defaults={
                                            "headcount":
                                                item[
                                                    "headcount"
                                                ],

                                            "productive_hours_per_day":
                                                Decimal(
                                                    "6.75"
                                                ),

                                            "working_days": 5,
                                        }
                                    )
                                )

                                if cap_created:
                                    created_capacities += 1
                                else:
                                    updated_capacities += 1

                        context.update({
                            "import_success": True,

                            "created_workcenters":
                                created_workcenters,

                            "created_capacities":
                                created_capacities,

                            "updated_capacities":
                                updated_capacities,
                        })

            except Exception as error:

                context.update({
                    "validation_failed": True,
                    "general_error": str(error),
                })

    return render(
        request,
        "capacity/capacity_import.html",
        context
    )


def download_capacity_template(request):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Ekip kapasitesi"

    headers = [
        "WorkCenter",
        "Adam sayı",
    ]

    for number, header in enumerate(
        headers,
        start=1
    ):
        cell = worksheet.cell(
            row=1,
            column=number,
            value=header
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    worksheet.append([
        "MEXANK-1",
        21
    ])

    worksheet.append([
        "QAYNAQ",
        20
    ])

    worksheet.append([
        "ELEKTK-1",
        16
    ])

    worksheet.column_dimensions[
        "A"
    ].width = 25

    worksheet.column_dimensions[
        "B"
    ].width = 18

    worksheet.freeze_panes = "A2"

    instruction = workbook.create_sheet(
        "Təlimat"
    )

    instruction.append([
        "MPS — Həftəlik Ekip Kapasitesi"
    ])

    instruction.append([])

    instruction.append([
        "WorkCenter",
        "SAP-da istifadə olunan ekip kodu."
    ])

    instruction.append([
        "Adam sayı",
        "Seçilmiş həftədə faktiki "
        "mövcud işçi sayı."
    ])

    instruction.append([])

    instruction.append([
        "Gündəlik kapasite",
        "Adam sayı × 6.75 MH"
    ])

    instruction.append([
        "Həftəlik kapasite",
        "Adam sayı × 6.75 × 5"
    ])

    instruction.column_dimensions[
        "A"
    ].width = 25

    instruction.column_dimensions[
        "B"
    ].width = 65

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="MPS_Capacity_Sablonu.xlsx"'
    )

    return response